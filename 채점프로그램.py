import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

class GradingApp:
    def __init__(self, master):
        self.master = master
        master.title("채점 프로그램")

        self.answer_key = []
        self.student_answers = []

        # 버튼 크기와 글씨 크기 조정
        button_width = 30
        button_height = 3
        button_font = ("Arial", 14)

        self.answer_key_button = tk.Button(master, text="정답 파일 업로드", command=self.upload_answer_key, width=button_width, height=button_height, font=button_font)
        self.answer_key_button.pack(pady=10)

        self.student_answers_button = tk.Button(master, text="학생 답안 파일 업로드", command=self.upload_student_answers, width=button_width, height=button_height, font=button_font)
        self.student_answers_button.pack(pady=10)

        self.grade_button = tk.Button(master, text="채점", command=self.grade, width=button_width, height=button_height, font=button_font)
        self.grade_button.pack(pady=10)

        self.result_label = tk.Label(master, text="")
        self.result_label.pack()

    def upload_answer_key(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        wb = load_workbook(file_path)
        ws = wb.active
        self.answer_key = [cell.value for cell in ws[2][1:]]  # 정답 행 읽기 (첫 번째 열 제외)

    def upload_student_answers(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        wb = load_workbook(file_path)
        ws = wb.active
        self.header = [cell.value for cell in ws[1]]  # 헤더 행 저장
        self.student_answers = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]  # 학생 답안 읽기

    def grade(self):
        scores = []
        wrong_answers = []
        for student in self.student_answers:
            score = 0
            wrong = []
            for i in range(len(self.answer_key)):
                if str(self.answer_key[i]) == str(student[i+2]):
                    score += 1
                else:
                    wrong.append((i+1, student[i+2], self.answer_key[i]))  # 틀린 문제 번호, 학생 답안, 정답 저장
            scores.append((student[0], student[1], score))  # 학번, 이름, 점수 튜플로 저장
            wrong_answers.append(wrong)  # 학생별 틀린 문제 정보 저장
        
        scores.sort(key=lambda x: x[2], reverse=True)  # 점수 기준으로 내림차순 정렬
        
        ranks = []
        prev_score = None
        prev_rank = 0
        for i, (student_id, name, score) in enumerate(scores, start=1):
            if score != prev_score:
                rank = i
                prev_rank = rank
            else:
                rank = prev_rank
            ranks.append((student_id, name, score, rank, wrong_answers[i-1]))  # 학번, 이름, 점수, 순위, 틀린 문제 정보 튜플로 저장
            prev_score = score
        
        # 결과를 Excel 파일로 저장
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        wb = Workbook()
        ws = wb.active
        ws.append(['학번', '이름', '점수', '순위', '틀린 문제(문제번호: 학생답안 -> 정답)'])
        for student_id, name, score, rank, wrong in ranks:
            ws.append([student_id, name, score, f'{rank}등' if rank == prev_rank else f'{rank}등', ', '.join(f'{num}: {student} -> {answer}' for num, student, answer in wrong)])  # 틀린 문제 정보를 문자열로 변환하여 추가
        wb.save(file_path)
        
        result = "채점 결과:\n"
        for student_id, name, score, rank, wrong in ranks:
            result += f"{name} ({student_id}): {score}점 (순위: {rank}등), 틀린 문제: {', '.join(f'{num}: {student} -> {answer}' for num, student, answer in wrong)}\n"
        
        self.result_label.config(text=result)

root = tk.Tk()
app = GradingApp(root)
root.mainloop()